import pandas as pd
from textblob import TextBlob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Step 1: Load dataset
input_file = "D:/twitter new sentiment/twitter_dataset.csv"
df = pd.read_csv(input_file)

# Step 2: Clean column names and print them
df.columns = df.columns.str.strip()
print("Available columns:", list(df.columns))  # Optional for debugging

# Step 3: Use correct column name for text
text_column = 'Text'  # <-- Corrected to match your dataset

# Step 4: Verify column exists
if text_column not in df.columns:
    raise KeyError(f"Column '{text_column}' not found. Available columns: {list(df.columns)}")

# Step 5: Define sentiment analysis function
def analyze_sentiment(text):
    if pd.isna(text):
        return 'Neutral', 0, 0.0

    blob = TextBlob(str(text))
    sentiment_score = blob.sentiment.polarity
    sentiment_percentage = abs(sentiment_score) * 100

    if sentiment_score > 0:
        sentiment = 'Positive'
    elif sentiment_score < 0:
        sentiment = 'Negative'
    else:
        sentiment = 'Neutral'

    return sentiment, sentiment_percentage, sentiment_score

# Step 6: Apply sentiment analysis
df[['Sentiment', 'Sentiment_Percentage', 'Sentiment_Score']] = df[text_column].apply(
    lambda x: pd.Series(analyze_sentiment(x))
)

# Step 7: Save to Excel
output_file = "D:/twitter new sentiment/twitter_dataset_colored.xlsx"
df.to_excel(output_file, index=False)

# Step 8: Open Excel workbook for formatting
wb = load_workbook(output_file)
ws = wb.active

# Step 9: Define color fills
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')   # Positive
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Negative

# Step 10: Find the column letter for "Sentiment"
sentiment_col_letter = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == 'Sentiment':
        sentiment_col_letter = col[0].column_letter
        break

if not sentiment_col_letter:
    raise ValueError("Could not find 'Sentiment' column in the Excel sheet.")

# Step 11: Apply color formatting row by row
for row in range(2, ws.max_row + 1):  # Skip header row
    cell = ws[f'{sentiment_col_letter}{row}']
    if cell.value == 'Positive':
        cell.fill = green_fill
    elif cell.value == 'Negative':
        cell.fill = yellow_fill
    # Neutral remains uncolored

# Step 12: Save formatted Excel file
wb.save(output_file)
print(f"✅ Sentiment analysis completed. File saved as: {output_file}")
